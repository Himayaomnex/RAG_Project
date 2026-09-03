"""
================================================================================
Daily Excel Rollup Generator (daily_excel_generator.py)
================================================================================
Generates production-grade, multi-tab Excel (.xlsx) workbooks summarizing daily
training progress, deliverables, learning gaps, and architectural decisions.

Pulls structured facts directly from Ganesh's Supabase Knowledge Base (kb schema)
and exports formatted spreadsheets to Google Drive / local deliverables directory.

Sheets:
1. Executive Summary & KPIs
2. Active Deliverables & Deadlines (kb.v_assignments_current)
3. Concept Gaps & Trainee Progress (kb.v_concepts)
4. Architectural Decisions Register (kb.v_decisions)
5. Mentor Feedback & QA Audit (kb.v_feedback & kb.v_qa)
================================================================================
"""

import os
import sys
import datetime
from typing import Optional, Dict, Any, List

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from agents.shared.kb_client import kb_client


# ── Color Palette & Styling Constants ──────────────────────────────────────────
NAVY_HEADER     = "1F4E79"  # Dark Corporate Navy
WHITE_TEXT      = "FFFFFF"
ACCENT_BLUE     = "2F5597"
ZEBRA_LIGHT     = "F2F4F7"
BORDER_GRAY     = "D9D9D9"
GREEN_FILL      = "E2EFDA"  # Completed / Delivered
GREEN_TEXT      = "375623"
YELLOW_FILL     = "FFF2CC"  # In Progress / Partial
YELLOW_TEXT     = "7F6000"
RED_FILL        = "FCE4D6"  # Late / Confused / Blocker
RED_TEXT        = "C65911"


def _apply_header_style(cell, text: str, bg_color: str = NAVY_HEADER, font_size: int = 11):
    cell.value = text
    cell.font = Font(name="Calibri", size=font_size, bold=True, color=WHITE_TEXT)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border():
    thin = Side(border_style="thin", color=BORDER_GRAY)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _auto_fit_columns(ws, max_len_cap: int = 60):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_length:
                max_length = len(val_str)
        adjusted_width = min(max(max_length + 4, 12), max_len_cap)
        ws.column_dimensions[col_letter].width = adjusted_width


def generate_daily_rollup_excel(
    target_date: Optional[str] = None,
    output_dir: Optional[str] = None,
    gdrive_dir: Optional[str] = None
) -> str:
    """
    Generates a multi-tab daily rollup workbook from the Knowledge Base and saves it.
    Returns the absolute file path of the generated Excel workbook.
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    date_str = target_date or datetime.date.today().strftime("%Y-%m-%d")
    wb = openpyxl.Workbook()

    # ── 1. SHEET 1: Executive Summary ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner
    ws1.merge_cells("A1:G1")
    title_cell = ws1["A1"]
    title_cell.value = f"Enterprise Multi-Agent Daily Rollup — {date_str}"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=WHITE_TEXT)
    title_cell.fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 35

    # Subtitle Metadata
    ws1.merge_cells("A2:G2")
    sub_cell = ws1["A2"]
    sub_cell.value = f"Data Source: Ganesh's PostgreSQL Knowledge Base (Supabase) | Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="595959")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    # Person State Table
    headers_ws1 = [
        "Team Member", "Assignments Open", "Delivered", "Late",
        "Concepts Demonstrated", "Concepts Confused/Partial", "Mentor Feedback"
    ]
    ws1.row_dimensions[4].height = 25
    for col_idx, h in enumerate(headers_ws1, 1):
        _apply_header_style(ws1.cell(row=4, column=col_idx), h, bg_color=ACCENT_BLUE)

    person_states = kb_client.get_person_state()
    row_idx = 5
    for p in person_states:
        # Dynamically exclude facilitators/mentors who have 0 trainee assignments and 0 feedback received
        if (p.get("assignments_open", 0) + p.get("assignments_delivered", 0) + p.get("feedback_received", 0)) == 0:
            continue
        ws1.cell(row=row_idx, column=1, value=p.get("person", "Unknown"))
        ws1.cell(row=row_idx, column=2, value=p.get("assignments_open", 0))
        ws1.cell(row=row_idx, column=3, value=p.get("assignments_delivered", 0))
        ws1.cell(row=row_idx, column=4, value=p.get("assignments_late", 0))
        ws1.cell(row=row_idx, column=5, value=p.get("concepts_demonstrated", 0))
        confused_count = (p.get("concepts_confused", 0) or 0) + (p.get("concepts_partial", 0) or 0)
        ws1.cell(row=row_idx, column=6, value=confused_count)
        ws1.cell(row=row_idx, column=7, value=p.get("feedback_received", 0))

        # Format row
        fill = PatternFill(start_color=ZEBRA_LIGHT if row_idx % 2 == 0 else "FFFFFF",
                           end_color=ZEBRA_LIGHT if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")
        for c in range(1, 8):
            cell = ws1.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.border = _thin_border()
            if c > 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(name="Calibri", bold=True)
        ws1.row_dimensions[row_idx].height = 22
        row_idx += 1

    # Totals Row
    ws1.cell(row=row_idx, column=1, value="Cohort Total")
    ws1.cell(row=row_idx, column=1).font = Font(name="Calibri", bold=True)
    ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 8):
        col_letter = get_column_letter(c)
        total_cell = ws1.cell(row=row_idx, column=c, value=f"=SUM({col_letter}5:{col_letter}{row_idx-1})")
        total_cell.font = Font(name="Calibri", bold=True)
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 8):
        ws1.cell(row=row_idx, column=c).border = _thin_border()
        ws1.cell(row=row_idx, column=c).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws1.row_dimensions[row_idx].height = 24
    _auto_fit_columns(ws1)

    # ── 2. SHEET 2: Deliverables & Assignments ─────────────────────────────────
    ws2 = wb.create_sheet(title="Deliverables & Tasks")
    ws2.views.sheetView[0].showGridLines = True
    headers_ws2 = ["Trainee", "Task Description", "Status", "Due Date", "Delivered Date", "Was Late?", "As of Session"]
    ws2.row_dimensions[1].height = 26
    for col_idx, h in enumerate(headers_ws2, 1):
        _apply_header_style(ws2.cell(row=1, column=col_idx), h)

    assignments = kb_client.get_current_assignments()
    for row_idx, a in enumerate(assignments, 2):
        status = (a.get("status") or "in_progress").lower()
        ws2.cell(row=row_idx, column=1, value=a.get("person", "Unknown"))
        ws2.cell(row=row_idx, column=2, value=a.get("task_description", ""))
        ws2.cell(row=row_idx, column=3, value=status.upper())
        ws2.cell(row=row_idx, column=4, value=str(a.get("due_date") or "Unspecified"))
        ws2.cell(row=row_idx, column=5, value=str(a.get("delivered_date") or "-"))
        ws2.cell(row=row_idx, column=6, value="YES" if a.get("was_late") else "NO")
        ws2.cell(row=row_idx, column=7, value=str(a.get("as_of_session") or "-"))

        for c in range(1, 8):
            cell = ws2.cell(row=row_idx, column=c)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", horizontal="left" if c in (1, 2) else "center")
            # Status Badge Styling
            if c == 3:
                if "delivered" in status or "completed" in status:
                    cell.fill = PatternFill(start_color=GREEN_FILL, fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, color=GREEN_TEXT)
                elif "progress" in status or "given" in status:
                    cell.fill = PatternFill(start_color=YELLOW_FILL, fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, color=YELLOW_TEXT)
                else:
                    cell.fill = PatternFill(start_color=RED_FILL, fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, color=RED_TEXT)
            else:
                cell.fill = PatternFill(start_color=ZEBRA_LIGHT if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")
        ws2.row_dimensions[row_idx].height = 20
    _auto_fit_columns(ws2)

    # ── 3. SHEET 3: Concept Gaps & Knowledge Tracking ──────────────────────────
    ws3 = wb.create_sheet(title="Concept Gaps & Gaps")
    ws3.views.sheetView[0].showGridLines = True
    headers_ws3 = ["Trainee", "Concept / Topic", "Understanding State", "Observation & Gap Notes", "Session Date"]
    ws3.row_dimensions[1].height = 26
    for col_idx, h in enumerate(headers_ws3, 1):
        _apply_header_style(ws3.cell(row=1, column=col_idx), h)

    concept_gaps = kb_client.get_concept_gaps()
    for row_idx, cg in enumerate(concept_gaps, 2):
        state = (cg.get("understanding_state") or "confused").lower()
        ws3.cell(row=row_idx, column=1, value=cg.get("person", "Unknown"))
        ws3.cell(row=row_idx, column=2, value=cg.get("concept", ""))
        ws3.cell(row=row_idx, column=3, value=state.upper())
        ws3.cell(row=row_idx, column=4, value=cg.get("observation", ""))
        ws3.cell(row=row_idx, column=5, value=str(cg.get("session_date") or "-"))

        for c in range(1, 6):
            cell = ws3.cell(row=row_idx, column=c)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", horizontal="left" if c in (1, 2, 4) else "center")
            if c == 3:
                if state == "confused":
                    cell.fill = PatternFill(start_color=RED_FILL, fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, color=RED_TEXT)
                else:
                    cell.fill = PatternFill(start_color=YELLOW_FILL, fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, color=YELLOW_TEXT)
            else:
                cell.fill = PatternFill(start_color=ZEBRA_LIGHT if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")
        ws3.row_dimensions[row_idx].height = 20
    _auto_fit_columns(ws3, max_len_cap=80)

    # ── 4. SHEET 4: Decisions Register ─────────────────────────────────────────
    ws4 = wb.create_sheet(title="Decisions Register")
    ws4.views.sheetView[0].showGridLines = True
    headers_ws4 = ["Session Date", "Owner", "Decision", "Rationale", "Scope"]
    ws4.row_dimensions[1].height = 26
    for col_idx, h in enumerate(headers_ws4, 1):
        _apply_header_style(ws4.cell(row=1, column=col_idx), h)

    decisions = kb_client.get_decisions()
    for row_idx, d in enumerate(decisions, 2):
        ws4.cell(row=row_idx, column=1, value=str(d.get("session_date") or "-"))
        ws4.cell(row=row_idx, column=2, value=d.get("owner") or "Team")
        ws4.cell(row=row_idx, column=3, value=d.get("decision_text", ""))
        ws4.cell(row=row_idx, column=4, value=d.get("rationale", ""))
        ws4.cell(row=row_idx, column=5, value=d.get("scope", "Project"))

        for c in range(1, 6):
            cell = ws4.cell(row=row_idx, column=c)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", horizontal="left" if c in (3, 4) else "center")
            cell.fill = PatternFill(start_color=ZEBRA_LIGHT if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")
        ws4.row_dimensions[row_idx].height = 20
    _auto_fit_columns(ws4, max_len_cap=80)

    # ── 5. File Persistence & Google Drive / OneDrive Sync ──────────────────────
    base_dir = output_dir or os.path.join(os.path.dirname(os.path.abspath(__file__)), "deliverables", "daily_rollups")
    os.makedirs(base_dir, exist_ok=True)
    filename = f"daily_rollup_{date_str}.xlsx"
    local_file_path = os.path.join(base_dir, filename)

    # Fail-safe save in case the file is currently open in Excel
    try:
        wb.save(local_file_path)
    except PermissionError:
        ts = datetime.datetime.now().strftime("%H%M%S")
        filename = f"daily_rollup_{date_str}_{ts}.xlsx"
        local_file_path = os.path.join(base_dir, filename)
        wb.save(local_file_path)
    print(f"[Excel Generator] Daily rollup saved locally: {local_file_path}")

    # If Cloud folder (OneDrive / Google Drive) path is configured, save a copy there too
    gdrive_path = gdrive_dir or os.getenv("GOOGLE_DRIVE_FOLDER")
    if gdrive_path:
        os.makedirs(gdrive_path, exist_ok=True)
        drive_file_path = os.path.join(gdrive_path, filename)
        try:
            wb.save(drive_file_path)
            print(f"[Excel Generator] Daily rollup synced to Cloud: {drive_file_path}")
        except PermissionError:
            ts = datetime.datetime.now().strftime("%H%M%S")
            drive_file_path = os.path.join(gdrive_path, f"daily_rollup_{date_str}_{ts}.xlsx")
            wb.save(drive_file_path)
            print(f"[Excel Generator] Daily rollup synced to Cloud: {drive_file_path}")

    # Direct Google Drive Cloud API Sync
    try:
        from gdrive_direct_uploader import upload_to_google_drive
        upload_to_google_drive(local_file_path)
    except Exception as e:
        print(f"[Google Drive API Hook] {e}")

    return local_file_path


if __name__ == "__main__":
    path = generate_daily_rollup_excel()
    print(f"Generated successfully at: {path}")
